import discord
import os
import openpyxl
from captcha.image import ImageCaptcha
import random
from discord import Member
import datetime
from discord.ext import commands
import asyncio
import youtube_dl
import re
import bs4
import time
import urllib
from urllib.request import urlopen, Request
from selenium import webdriver

client = discord.Client()

@client.event
async def on_ready():
    print("login")
    print(client.user.name)
    print(client.user.id)
    print("-------------------")
    print("ready")
    game = discord.Game("🍊코로나에는 귤이 좋데요!🍊")
    await client.change_presence(status=discord.Status.online, activity=game)


@client.event
async def on_message(message):

    if message.content.startswith("안녕하세요"):
        msg = '🌲아, 드디어 오셨군요, {0.author.mention}님!🌳'.format(message)
        await message.channel.send(msg)

    if message.content.startswith("!26/배그하자"):
        randomNum = random.randrange(1, 3)
        if randomNum==1:
            await message.channel.send(embed=discord.Embed(title="네! 배그합시다!", color=discord.Color.blue()))
        else:
            await message.channel.send(embed=discord.Embed(title="자러갑시다....", color=discord.Color.red()))

    if message.content.startswith("!26/홋치"):
        file = openpyxl.load_workbook('기억.xlsx')
        sheet = file.active
        learn = message.content.split(" ")
        for i in range(1, 201):
            if sheet["A"+str(i)].value == "-":
                sheet["A" + str(i)].value = learn[1]
                sheet["B" + str(i)].value = learn[2]
                await message.channel.send("좀더 때려줘어어엉항항하ㅑㅎ아아")
                await message.channel.send("★ 현재 사용중인 데이터 저장용량 : 200/" + str(i)+" ★")
                break
        file.save("기억.xlsx")

    if message.content.startswith("!26/말해"):
        file = openpyxl.load_workbook("기억.xlsx")
        sheet = file.active
        memory = message.content.split(" ")
        for i in range(1, 201):
            if sheet["A" + str(i)].value == memory[1]:
                await message.channel.send(sheet["B" + str(i)].value)
                break

    if message.content.startswith("!26/기억 초기화") or message.content.startswith("!26/기억초기화"):
        file = openpyxl.load_workbook("기억.xlsx")
        sheet = file.active
        for i in range(1, 251):
            sheet["A"+str(i)].value = "-"
        await message.channel.send("기억초기화 완료")
        file.save("기억.xlsx")

    if message.content.startswith("!26/데이터목록") or message.content.startswith("!26/데이터 목록"):
        file = openpyxl.load_workbook("기억.xlsx")
        sheet = file.active
        for i in range(1, 201):
            if sheet["A" + str(i)].value == "-" and i == 1:
                await message.channel.send("데이터 없음")
            if sheet["A" + str(i)].value == "-":
                break
            await message.channel.send("A : "+sheet["A" + str(i)].value + " B : "+ sheet["B" + str(i)].value)

    if message.content.startswith("!26/모두모여"):
        await message.channel.send("@everyone")

    if message.content.startswith("!26/여기모여"):
        await message.channel.send("@here")

    if message.content.startswith("!26/날씨"):
        learn = message.content.split(" ")
        enc_location = urllib.parse.quote('날씨')
        hdr = {'User-Agent': 'Mozilla/5.0'}
        url = 'https://search.naver.com/search.naver?where=nexearch&sm=top_hty&fbm=1&ie=utf8&query=' + enc_location
        print(url)
        req = Request(url, headers=hdr)
        html = urllib.request.urlopen(req)
        bsObj = bs4.BeautifulSoup(html, "html.parser")
        todayBase = bsObj.find('div', {'class': 'main_info'})

        todayTemp1 = todayBase.find('span', {'class': 'todaytemp'})
        todayTemp = todayTemp1.text.strip()  # 온도
        print(todayTemp)

        todayValueBase = todayBase.find('ul', {'class': 'info_list'})
        todayValue2 = todayValueBase.find('p', {'class': 'cast_txt'})
        todayValue = todayValue2.text.strip()  # 밝음,어제보다 ?도 높거나 낮음을 나타내줌
        print(todayValue)

        todayFeelingTemp1 = todayValueBase.find('span', {'class': 'sensible'})
        todayFeelingTemp = todayFeelingTemp1.text.strip()  # 체감온도
        print(todayFeelingTemp)

        todayMiseaMongi1 = bsObj.find('div', {'class': 'sub_info'})
        todayMiseaMongi2 = todayMiseaMongi1.find('div', {'class': 'detail_box'})
        todayMiseaMongi3 = todayMiseaMongi2.find('dd')
        todayMiseaMongi = todayMiseaMongi3.text  # 미세먼지
        print(todayMiseaMongi)

        tomorrowBase = bsObj.find('div', {'class': 'table_info weekly _weeklyWeather'})
        tomorrowTemp1 = tomorrowBase.find('li', {'class': 'date_info'})
        tomorrowTemp2 = tomorrowTemp1.find('dl')
        tomorrowTemp3 = tomorrowTemp2.find('dd')
        tomorrowTemp = tomorrowTemp3.text.strip()  # 오늘 오전,오후온도
        print(tomorrowTemp)

        tomorrowAreaBase = bsObj.find('div', {'class': 'tomorrow_area'})
        tomorrowMoring1 = tomorrowAreaBase.find('div', {'class': 'main_info morning_box'})
        tomorrowMoring2 = tomorrowMoring1.find('span', {'class': 'todaytemp'})
        tomorrowMoring = tomorrowMoring2.text.strip()  # 내일 오전 온도
        print(tomorrowMoring)

        tomorrowValue1 = tomorrowMoring1.find('div', {'class': 'info_data'})
        tomorrowValue = tomorrowValue1.text.strip()  # 내일 오전 날씨상태, 미세먼지 상태
        print(tomorrowValue)

        tomorrowAreaBase = bsObj.find('div', {'class': 'tomorrow_area'})
        tomorrowAllFind = tomorrowAreaBase.find_all('div', {'class': 'main_info morning_box'})
        tomorrowAfter1 = tomorrowAllFind[1]
        tomorrowAfter2 = tomorrowAfter1.find('p', {'class': 'info_temperature'})
        tomorrowAfter3 = tomorrowAfter2.find('span', {'class': 'todaytemp'})
        tomorrowAfterTemp = tomorrowAfter3.text.strip()  # 내일 오후 온도
        print(tomorrowAfterTemp)

        tomorrowAfterValue1 = tomorrowAfter1.find('div', {'class': 'info_data'})
        tomorrowAfterValue = tomorrowAfterValue1.text.strip()

        print(tomorrowAfterValue)  # 내일 오후 날씨상태,미세먼지

        embed = discord.Embed(
            title="주인님 지역의 날씨 정보".format(message),
            description='날씨 정보입니다.',
            colour=discord.Colour.gold()
        )
        embed.add_field(name='현재온도', value=todayTemp+'˚', inline=False)  # 현재온도
        embed.add_field(name='체감온도', value=todayFeelingTemp, inline=False)  # 체감온도
        embed.add_field(name='현재상태', value=todayValue, inline=False)  # 밝음,어제보다 ?도 높거나 낮음을 나타내줌
        embed.add_field(name='현재 미세먼지 상태', value=todayMiseaMongi, inline=False)  # 오늘 미세먼지
        embed.add_field(name='오늘 오전/오후 날씨', value=tomorrowTemp, inline=False)  # 오늘날씨 # color=discord.Color.blue()
        embed.add_field(name='**----------------------------------**',value='**----------------------------------**', inline=False)  # 구분선
        embed.add_field(name='내일 오전온도', value=tomorrowMoring+'˚', inline=False)  # 내일오전날씨
        embed.add_field(name='내일 오전날씨상태, 미세먼지 상태', value=tomorrowValue, inline=False)  # 내일오전 날씨상태
        embed.add_field(name='내일 오후온도', value=tomorrowAfterTemp + '˚', inline=False)  # 내일오후날씨
        embed.add_field(name='내일 오후날씨상태, 미세먼지 상태', value=tomorrowAfterValue, inline=False)  # 내일오후 날씨상태
        await message.channel.send(embed=embed)

    if message.content.startswith("!26/롤"):
        learn = message.content.split(" ")
        location = learn[1]
        enc_location = urllib.parse.quote(location)

        url = "http://www.op.gg/summoner/userName=" + enc_location
        html = urllib.request.urlopen(url)

        bsObj = bs4.BeautifulSoup(html, "html.parser")
        rank1 = bsObj.find("div", {"class": "TierRankInfo"})
        rank2 = rank1.find("div", {"class": "TierRank"})
        rank3 = rank2.find("span", {"class": "tierRank"})
        rank4 = rank3.text  # 티어표시 (브론즈1,2,3,4,5 등등)
        print(rank4)
        if rank4 != 'Unranked':
          jumsu1 = rank1.find("div", {"class": "TierInfo"})
          jumsu2 = jumsu1.find("span", {"class": "LeaguePoints"})
          jumsu3 = jumsu2.text
          jumsu4 = jumsu3.strip()#점수표시 (11LP등등)
          print(jumsu4)

          winlose1 = jumsu1.find("span", {"class": "WinLose"})
          winlose2 = winlose1.find("span", {"class": "wins"})
          winlose2_1 = winlose1.find("span", {"class": "losses"})
          winlose2_2 = winlose1.find("span", {"class": "winratio"})

          winlose2txt = winlose2.text
          winlose2_1txt = winlose2_1.text
          winlose2_2txt = winlose2_2.text #승,패,승률 나타냄  200W 150L Win Ratio 55% 등등

          print(winlose2txt + " " + winlose2_1txt + " " + winlose2_2txt)

        channel = message.channel
        embed = discord.Embed(
            title='롤 정보',
            description='롤 정보입니다.',
            colour=discord.Colour.green()
        )
        if rank4=='Unranked':
            embed.add_field(name='당신의 티어', value=rank4, inline=False)
            embed.add_field(name='-당신은 언랭-', value="언랭은 더이상의 정보는 제공하지 않습니다.", inline=False)
            await message.channel.send(embed=embed)
        else:
         embed.add_field(name='당신의 티어', value=rank4, inline=False)
         embed.add_field(name='당신의 LP(점수)', value=jumsu4, inline=False)
         embed.add_field(name='당신의 승,패 정보', value=winlose2txt+" "+winlose2_1txt, inline=False)
         embed.add_field(name='당신의 승률', value=winlose2_2txt, inline=False)
         await message.channel.send(embed=embed)


    if message.content.startswith("!26/업데이트"):
        msg = '{0.author.mention} 🤳이번주 업데이트 내역입니다(v 1.3).🤳 ```1️⃣안녕하세요 를 하면 맨션으로 자동 반응이 됩니다!``` ```2️⃣경고가 안 된다는 오류를 고쳤습니다.```'.format(message)
        await message.channel.send(msg)

    if message.content.startswith("!봇사진"):
        pic = message.content.split(" ")[1]
        await message.channel.send(file=discord.File(pic))

    if message.content.startswith("!채널메세지"):
        channel = message.content[7:25]
        msg = message.content[26:]
        msg2 = '{0.author.mention} 님이 보내셨습니다. '.format(message)
        await client.get_channel(int(channel)).send(msg2 + msg)

    if message.content.startswith("!DM"):
        author = message.guild.get_member(int(message.content[4:22]))
        msg = message.content[23:]
        await author.send(msg)

    if message.content.startswith("!뮤트"):
        author = message.guild.get_member(int(message.content[4:22]))
        role = discord.utils.get(message.guild.roles, name="뮤트")
        await author.add_roles(role)
        await message.channel.send("⚠️신고를 받아서 출동하였습니다. 다음에는 좀더 좋게 생활하는 것도 좋을 것 같아요⚠️")

    if message.content.startswith("!언뮤트"):
        author = message.guild.get_member(int(message.content[5:23]))
        role = discord.utils.get(message.guild.roles, name="뮤트")
        await author.remove_roles(role)
        await message.channel.send("❤️다음에는 실수 하지 않게 조심하세요!❤️")

    if message.content.startswith("!26/병신"):
        await message.channel.send("제가 병신인 이유가 뭐죠? 도대체 왜죠?")
        time.sleep(5)
        await message.channel.send("그 마땅한 이유가 있다면 말해주세요.")
        time.sleep(3)
        await message.author.send("왜죠? 왜죠? 왜죠? 왜죠? 왜죠?")
        await message.author.send("왜죠? 왜죠? 왜죠? 왜죠? 왜죠? 왜죠? 왜죠? 왜죠? 왜죠? 왜죠? 왜죠? 왜죠? 왜죠? 왜죠? 왜죠? 왜죠? 왜죠? 왜죠? 왜죠? 왜죠? 왜죠? 왜죠? 왜죠? 왜죠? 왜죠?")
        await message.author.send("내가 얼마나 널 챙겨줬는데.")
        await message.author.send("Y̷̨̤̮̦̭͇̩̾͆̈́̍̀͊̓́͘̕͜O̷̳̙̞͉̮̫͂͐́Ư̵̢͔̽̏̊̕͝ ̵̢̗̗̺͉͎̘̗̘͎̈͆D̸͇̰̠̩͊͛̃͂̎͋̕͝O̸̢͎̞̖̮̬͇̳̙̘͋̌͑̕Ṋ̶̛͗̇̆̾̾̕̚'̵̮̦̭͇͌͐́̚͠T̵͚̝͙̾͑̓͜ ̶͇͂͑͌̇͝Ḩ̴̡͓̫̹̝̤͗Ā̵̭̃V̶̹͒͆E̷̢̢̪͎̦̦͖̳̭̓͑ ̵̛̜̜͍͈͎̿̇T̴̼̜͈̠͉͙͒͜Ǫ̴̻̘̬̝̦͋͗̀̑̽́͌̚͠ ̴̨̥̤̩͒̾͆̇̎̋Ḑ̸̢̖̲̖̟͋͊̓E̶̛̛̞͐̍͑̂̇̔͗͝S̴̤͍̫̫͈͇͈̹̭̫̊̀É̸̘̠͙͖̣̣̻̮̰͑́̐̍͗͌̿ͅR̵̺̱̍̏̑͆V̶͉͋͒̎̅̆̀͒̉̄̌Ė̷̢̡̛̮̜̭̦̾̿̆́̐́̂͒ ̷͔͚͙̝̝̜̈́̋̓̽̿͌̚Ą̷͇͎̱̲̞͔̣͍̔̍ͅN̵̜̘̟̦̗̒̃͂̑̀̊̔̕͝Y̶̻̟̯̩̠̬̖̓͌̏̏́͝T̶͙̮̥͍͚̩͖͇̽̄̔̇̌͜͝Ḧ̸͎̳̦̫̞́͐̈̓͛̋̄̉Ḯ̴̻̤̪͗͑̆̿̽͊̈͝N̷̡̤̣̥̫͇̪̿̈́G̸̟͍̝̿̈́̎̔")

    if message.content.startswith("!26/배그듀오"):

        learn = message.content.split(" ")
        location = learn[1]
        enc_location = urllib.parse.quote(location)
        url = "https://dak.gg/profile/" + enc_location
        html = urllib.request.urlopen(url)
        bsObj = bs4.BeautifulSoup(html, "html.parser")
        duoCenter1 = bsObj.find("section", {"class": "duo modeItem"})
        duoRecord1 = duoCenter1.find("div", {"class": "overview"})
        duoRecord = duoRecord1.text.strip()  # ----기록이없습니다 문구----
        print(duoRecord)
        channel = message.channel
        embed = discord.Embed(
            title='배그듀오 정보',
            description='배그듀오 정보입니다.',
            colour=discord.Colour.green())
        if duoRecord == 'No record':
            print('듀오 경기가 없습니다.')
            embed.add_field(name='배그를 한판이라도 해주세요', value='듀오 경기 전적이 없습니다..', inline=False)
            await message.channel.send(embed=embed)

        else:
            duoRat1 = duoRecord1.find("span", {"class": "value"})
            duoRat = duoRat1.text.strip()  # ----레이팅----
            duoRank1 = duoRecord1.find("p", {"class": "grade-name"})
            duoRank = duoRank1.text.strip()  # ----등급----
            print(duoRank)
            embed.add_field(name='레이팅', value=duoRat, inline=False)
            embed.add_field(name='등급', value=duoRank, inline=False)


            duoStat = duoCenter1.find("div", {"class": "stats"})

            duoKD1 = duoStat.find("div", {"class": "kd stats-item stats-top-graph"})
            duoKD2 = duoKD1.find("p", {"class": "value"})
            duoKD = duoKD2.text.strip()  # ----킬뎃----
            duoKdSky1 = duoStat.find("span", {"class": "top"})
            duoKdSky = duoKdSky1.text.strip()  # ----킬뎃 상위?%----
            print(duoKD)
            print(duoKdSky)
            embed.add_field(name='킬뎃,킬뎃상위', value=duoKD+" "+duoKdSky, inline=False)

            duoWinRat1 = duoStat.find("div", {"class": "winratio stats-item stats-top-graph"})
            duoWinRat2 = duoWinRat1.find("p", {"class": "value"})
            duoWinRat = duoWinRat2.text.strip()  # ----승률----
            duoWinRatSky1 = duoWinRat1.find("span", {"class": "top"})
            duoWinRatSky = duoWinRatSky1.text.strip()  # ----승률 상위?%----
            print(duoWinRat)
            print(duoWinRatSky)
            embed.add_field(name='승률,승률상위', value=duoWinRat + " " + duoWinRatSky, inline=False)

            duoHead1 = duoStat.find("div", {"class": "headshots"})
            duoHead2 = duoHead1.find("p", {"class": "value"})
            duoHead = duoHead2.text.strip()  # ----헤드샷----
            duoHeadSky1 = duoHead1.find("span", {"class": "top"})
            duoHeadSky = duoHeadSky1.text.strip()  # ----헤드샷 상위?%----
            print(duoHead)
            print(duoHeadSky)
            embed.add_field(name='헤드샷,헤드샷상위', value=duoHead + " " + duoHeadSky, inline=False)
            await message.channel.send(embed=embed)


    if message.content.startswith('!경고부여') :
        author = message.guild.get_member(int(message.content[9:27]))
        file = openpyxl.load_workbook('경고.xlsx')
        sheet = file.active
        why = str(message.content[28:])
        i = 1
        while True :
            if sheet["A" + str(i)].value == str(author) :
                sheet['B' + str(i)].value = int(sheet["B" + str(i)].value) + 1
                file.save("경고.xlsx")
                if sheet["B" + str(i)].value == 4:
                    await message.guild.ban(author)
                    await message.channel.send(str(author) + "님은 경고 4회누적으로 서버에서 추방되었습니다.")
                else:
                    await message.channel.send(str(author) + "님은 경고를 1회 받았습니다")
                    sheet["c" + str(i)].value = why
                break
            if sheet["A" + str(i)].value == None:
                sheet["A" + str(i)].value = str(author)
                sheet["B" + str(i)].value = 1
                sheet["c" + str(i)].value = why
                file.save("경고.xlsx")
                await message.channel.send(str(author) + "님은 경고를 1회 받았습니다.")
                break
            i += 1

    if message.content.startswith("!26/배그스쿼드"):

        learn = message.content.split(" ")
        location = learn[1]
        enc_location = urllib.parse.quote(location)
        url = "https://dak.gg/profile/" + enc_location
        html = urllib.request.urlopen(url)
        bsObj = bs4.BeautifulSoup(html, "html.parser")
        duoCenter1 = bsObj.find("section", {"class": "squad modeItem"})
        duoRecord1 = duoCenter1.find("div", {"class": "overview"})
        duoRecord = duoRecord1.text.strip()  # ----기록이없습니다 문구----
        print(duoRecord)
        channel = message.channel
        embed = discord.Embed(
            title='배그스쿼드 정보',
            description='배그스쿼드 정보입니다.',
            colour=discord.Colour.green())
        if duoRecord == 'No record':
            print('스쿼드 경기가 없습니다.')
            embed.add_field(name='배그를 한판이라도 해주세요', value='스쿼드 경기 전적이 없습니다..', inline=False)
            await message.channel.send(embed=embed)

        else:
            duoRat1 = duoRecord1.find("span", {"class": "value"})
            duoRat = duoRat1.text.strip()  # ----레이팅----
            duoRank1 = duoRecord1.find("p", {"class": "grade-name"})
            duoRank = duoRank1.text.strip()  # ----등급----
            print(duoRank)
            embed.add_field(name='레이팅', value=duoRat, inline=False)
            embed.add_field(name='등급', value=duoRank, inline=False)


            duoStat = duoCenter1.find("div", {"class": "stats"})

            duoKD1 = duoStat.find("div", {"class": "kd stats-item stats-top-graph"})
            duoKD2 = duoKD1.find("p", {"class": "value"})
            duoKD = duoKD2.text.strip()  # ----킬뎃----
            duoKdSky1 = duoStat.find("span", {"class": "top"})
            duoKdSky = duoKdSky1.text.strip()  # ----킬뎃 상위?%----
            print(duoKD)
            print(duoKdSky)
            embed.add_field(name='킬뎃,킬뎃상위', value=duoKD+" "+duoKdSky, inline=False)

            duoWinRat1 = duoStat.find("div", {"class": "winratio stats-item stats-top-graph"})
            duoWinRat2 = duoWinRat1.find("p", {"class": "value"})
            duoWinRat = duoWinRat2.text.strip()  # ----승률----
            duoWinRatSky1 = duoWinRat1.find("span", {"class": "top"})
            duoWinRatSky = duoWinRatSky1.text.strip()  # ----승률 상위?%----
            print(duoWinRat)
            print(duoWinRatSky)
            embed.add_field(name='승률,승률상위', value=duoWinRat + " " + duoWinRatSky, inline=False)

            duoHead1 = duoStat.find("div", {"class": "headshots"})
            duoHead2 = duoHead1.find("p", {"class": "value"})
            duoHead = duoHead2.text.strip()  # ----헤드샷----
            duoHeadSky1 = duoHead1.find("span", {"class": "top"})
            duoHeadSky = duoHeadSky1.text.strip()  # ----헤드샷 상위?%----
            print(duoHead)
            print(duoHeadSky)
            embed.add_field(name='헤드샷,헤드샷상위', value=duoHead + " " + duoHeadSky, inline=False)
            await message.channel.send(embed=embed)

    if message.content.startswith('!26/고양이'):
        embed = discord.Embed(
            title='고양이는',
            description='멍멍',
            colour=discord.Colour.green()
        )

        urlBase = 'https://loremflickr.com/320/240?lock='
        randomNum = random.randrange(1, 30977)
        urlF = urlBase+str(randomNum)
        embed.set_image(url = urlF)
        await message.channel.send(embed=embed)

    if message.content.startswith('!강아지'):
        embed = discord.Embed(
            title='강아지는',
            description='야옹야옹',
            colour=discord.Colour.green()
        )

        urlBase = 'https://loremflickr.com/320/240/dog?lock='
        randomNum = random.randrange(1, 30977)
        urlF = urlBase+str(randomNum)
        embed.set_image(url = urlF)
        await message.channel.send(embed=embed)

    if message.content.startswith('!네코'):
        embed = discord.Embed(
            colour=discord.Colour.green()
        )
        embed2 = discord.Embed(
            colour=discord.Colour.green()
        )
        embed3 = discord.Embed(
            colour=discord.Colour.green()
        )
        randomnumber = random.randrange(100, 407)
        randomgiho = random.randrange(1,3)
        print('?번째사진 : '+str(randomnumber))
        print('기호 : '+str(randomgiho))
        strandomnumber = str(randomnumber)
        file1 = '.png'
        file2 = '.jpg'
        file3 = '.jpeg'
        giho = '_'
        if randomgiho==1:
            urlbase1 = "https://cdn.nekos.life/neko/neko" + strandomnumber + file1
            urlbase2 = "https://cdn.nekos.life/neko/neko" + strandomnumber + file2
            urlbase3 = "https://cdn.nekos.life/neko/neko" + strandomnumber + file3
            embed.set_image(url=urlbase1)
            embed2.set_image(url=urlbase2)
            embed3.set_image(url=urlbase3)
            await message.channel.send(embed=embed)
            await message.channel.send(embed=embed2)
            await message.channel.send(embed=embed3)
        else:
            urlbase_1 = "https://cdn.nekos.life/neko/neko" + giho + strandomnumber + file1
            urlbase_2 = "https://cdn.nekos.life/neko/neko" + giho + strandomnumber + file2
            urlbase_3 = "https://cdn.nekos.life/neko/neko" + giho + strandomnumber + file3
            embed.set_image(url=urlbase_1)
            embed2.set_image(url=urlbase_2)
            embed3.set_image(url=urlbase_3)
            await message.channel.send(embed=embed)
            await message.channel.send(embed=embed2)
            await message.channel.send(embed=embed3)

    if message.content.startswith('!실시간검색어') or message.content.startswith('!실검'):
        await message.channel.send("네이버 실시간 검색어가 폐지된 관계로 실시간 검색어 명령어는 비활성화 되었습니다.")

        embed = discord.Embed(
            title='네이버 실시간 검색어',
            description='실시간검색어',
            colour=discord.Colour.green()
        )
        for i in range(0,20):
            realTimeSerach4 = realTimeSerach3[i]
            realTimeSerach5 = realTimeSerach4.find('span', {'class': 'ah_k'})
            realTimeSerach = realTimeSerach5.text.replace(' ', '')
            realURL = 'https://search.naver.com/search.naver?ie=utf8&query='+realTimeSerach
            print(realTimeSerach)
            embed.add_field(name=str(i+1)+'위', value='\n'+'[%s](<%s>)' % (realTimeSerach, realURL), inline=False) # [텍스트](<링크>) 형식으로 적으면 텍스트 하이퍼링크 만들어집니다

        await message.channel.send(embed=embed)

    if message.content.startswith("!캡챠"):
        Image_captcha = ImageCaptcha()
        msg = ""
        a = ""
        for i in range(6):
            a += str(random.randint(0, 9))

        name = str(message.author.id) + ".png"
        Image_captcha.write(a, name)

        await message.channel.send(file=discord.File(name))
        def check(msg):
            return msg.author == message.author and msg.channel == message.channel

        try:
            msg = await client.wait_for("message", timeout=10, check=check)
        except:
            await message.channel.send("⏰시간 초과하셨어요!⏰")
            return

        if msg.content == a:
            await message.channel.send("☑️맞으셨어요!☑️")
        else:
            await message.channel.send("❌오답입니다ㅠ❌")

    if message.content.startswith("!26/쓰레기"):
        await message.channel.send("😨뭐, 뭐라고요?😨 저기 주인니..니,, 님,,ㅁ :warning: 26봇 에러 생겼습니다. !26/고치기를 하여 고치세요. :warning:")

    if message.content.startswith("!26/고치기"):
        await message.channel.send(":thumbsup: 에러가 고쳐졌습니다. :thumbsup:")

    if message.content.startswith("!26/서버등록"):
        await message.channel.send("👋안녕하세요. 여러분들의 서버의 맞배너를 받아 줄 26봇이라고 합니다.👋 ```등록을 할려면 등록할래 라고 말해주세요. 취소하고 싶다면 취소 라고 말해주세요. 이 작업은 중간에 취소를 언제든지 할 수 있습니다.```")

    if message.content.startswith("등록할래"):
        await message.channel.send("👌알겠습니다. 좋은 생각입니다! 지금 바로 등록 하겠습니다.👌 ``등록 하실 거면 좋아 이라고 말해주세요. 싫으시면 하지마 라고 말해주세요.``")

    if message.content.startswith("취소"):
        await message.channel.send("👌알겠습니다. 주인님 마음대로죠👌")

    if message.content.startswith("하지마"):
        await message.channel.send("😀실행 취소하겠습니다!😀")\

    if message.content.startswith("좋아"):
        await message.channel.send("😀그러면 𝟚𝟞𝕛𝕡𝟘𝟛. δ#8277 한테 친추를 보내시고 DM 보내주세요!😀")

    if message.content.startswith("!26MJ/마주정보"):
        await message.channel.send("🤖마주#5903 의 정보 ```이름: 마주``` ```트위치: blueblackslime( https://www.twitch.tv/blueblackslime )``` ```유튜브: 마주 ( https://www.youtube.com/channel/UCXeXeCn960Hm-y289HCAxkQ )``` ")

    if message.content.startswith("!시간"):
        a = datetime.datetime.today().year
        b = datetime.datetime.today().month
        c = datetime.datetime.today().day
        d = datetime.datetime.today().hour
        e = datetime.datetime.today().minute
        await message.channel.send(str(a) + "년 " + str(b) + "월 " + str(c) + "일 " + str(d) + "시 " + str(e) + "분 입니다!")

    if message.content.startswith("!좀비탄"):
        await message.channel.send("똑똑하고 영리한 너구리 아닌가용~")

    if message.content.startswith("!스베"):
        await message.channel.send("멍청이 아닌가요...?")

    if message.content.startswith("!마주"):
        await message.channel.send("와... 말할것 없이 몸매 섹시하고 개씹존잘 인성 천사 킹갓짱마주 아니겠습니까!")

    if message.content.startswith("!하노이"):
        await message.channel.send("뭐, 뭐라고? 반동이다! 아아ㅏ아아ㅏ!!!! https://im6.ezgif.com/tmp/ezgif-6-4b4007d8f851.gif")

    if message.content.startswith("!26/서버리스트"):
        await message.channel.send("👀등록된 서버들👀 ```👏대한민국 로블록스 게임중에서 가장 오래된 화랑부대 게임, 🛡화랑부대로 오세요!🛡 서버링크: https://discord.gg/nWxY5bV```")

    if message.content.startswith("!26/프로필등록"):
        await message.channel.send("⚠️정말이죠? 프로필을 등록하면 신고가 될 수 있습니다.⚠️ ``!계속`` 아니면 ``!그냥취소``")

    if message.content.startswith("!계속"):
        await message.channel.send("❤️넵, 잠시만 기다려주세요. 𝟚𝟞냥#8277 태그 해주세요!❤️")

    if message.content.startswith("!그냥취소"):
        await message.channel.send("❤️넵, 취소되었습니다.❤️")

    if message.content.startswith("!26/로그인"):
        await message.channel.send("🔐코드를 입력 부탁드립니다. ``다른 분이 사용시 처벌이 됩니다.``🔐")

    if message.content.startswith("!26/4263"):
        await message.channel.send("🎉어서오세요, 시리님!🎉")

    if message.content.startswith("!26/8277"):
        await message.channel.send("🎉어서오세요, 26님! 아! 맞다 하노이님이죠!🎉")

    if message.content.startswith("와! 샌즈!"):
        embed = discord.Embed(title="겁나 어.렵.습.니.다", description="언더테일 아시는구나! 혹시 모르시는분들에 대해 설명해드립니다. 샌즈랑 언더테일의 세가지 엔딩루트중 몰살엔딩의 최종보스로 진.짜.겁.나.어.렵.습.니.다. 공격은 전부다 회피하고 만피가 92인데 샌즈의 공격은 1초당 60이 다는데다가 독뎀까지 추가로 붙어있습니다.. 하지만 이러면 절대로 게임을 깰 수 가 없으니 제작진이 치명적인 약점을 만들었죠. 샌즈의 치명적인 약점이 바로 지친다는것입니다. 패턴들을 다 견디고나면 지쳐서 자신의 턴을 유지한채로 잠에듭니다. 하지만 잠이들었을때 창을 옮겨서 공격을 시도하고 샌즈는 1차공격은 피하지만 그 후에 바로날아오는 2차 공격을 맞고 죽습니다.", color=0x00B992)
        embed.set_image(url="https://image.librewiki.net/f/f0/Sans.gif")
        embed.set_thumbnail(url="https://pbs.twimg.com/profile_images/693527349254422528/51LpoPnZ_400x400.png ")
        await message.channel.send("와! 샌즈 아시는구나!", embed=embed)

    if message.content.startswith('!번역'):
        learn = message.content.split(" ")
        Text = ""

        client_id = ""
        client_secret = ""

        url = "https://openapi.naver.com/v1/papago/n2mt"
        print(len(learn))
        vrsize = len(learn)  # 배열크기
        vrsize = int(vrsize)
        for i in range(1, vrsize): #띄어쓰기 한 텍스트들 인식함
            Text = Text+" "+learn[i]
        encText = urllib.parse.quote(Text)
        data = "source=ko&target=en&text=" + encText

        request = urllib.request.Request(url)
        request.add_header("X-Naver-Client-Id", client_id)
        request.add_header("X-Naver-Client-Secret", client_secret)

        response = urllib.request.urlopen(request, data=data.encode("utf-8"))

        rescode = response.getcode()
        if (rescode == 200):
            response_body = response.read()
            data = response_body.decode('utf-8')
            data = json.loads(data)
            tranText = data['message']['result']['translatedText']
        else:
            print("Error Code:" + rescode)

        print('번역된 내용 :', tranText)

        embed = discord.Embed(
            title='한글->영어 번역결과',
            description=tranText,
            colour=discord.Colour.green()
        )
        await message.channel.send(embed=embed)

        
    if message.content.startswith('!움짤'):
        embed = discord.Embed(
            title='랜덤움짤',
            description='쨜쨜쨜쨜ㅉ랴ㅉ랴ㅉ랴ㅉ랴ㅉ랴ㅉ랴ㅉ랼',
            colour=discord.Colour.green()
        )
        url = "http://www.gifbin.com/random"
        urlBase = "http://www.gifbin.com"
        html = urllib.request.urlopen(url)
        bsObj = bs4.BeautifulSoup(html, "html.parser")
        gif1 = bsObj.find('form', {'id': 'share-form'})
        gif2 = gif1.find('a')
        gif3 = gif2["href"]
        gifURL = urlBase + gif3
        print(gifURL)
        embed.set_image(url=gifURL)
        await message.channel.send(embed=embed)
        #"http://www.gifbin.com/random"


    if message.content.startswith("!26/명령어"):
        embed = discord.Embed(title="명령어목록", description="더 많은 명령어가 추가됩니다!", color=0x00B992)
        embed.add_field(name="!봇사진", value="봇 파일에 등록된 사진을 가져옵니다.", inline=False)
        embed.add_field(name="!채널메세지", value="채널에 봇이 메세지를 보냅니다.", inline=False)
        embed.add_field(name="!DM", value="특정한 분한테 DM을 보냅니다.", inline=False)
        embed.add_field(name="!캡챠", value="로봇 인증을 테스트 하는 캡챠 시스템을 꺼냅니다.", inline=False)
        embed.add_field(name="!26/쓰레기", value="26이한테 약한 욕을 해서 고장네 트립니다.(비추천)", inline=False)
        embed.add_field(name="!26/자기소개", value="26이가 자기소개롤 하게 하는 명령어입니다.", inline=False)
        embed.add_field(name="!26/프로필등록", value="26이가 프로필을 등록해줍니다.", inline=False)
        embed.add_field(name="!시간", value="26이가 시간을 아르켜 주는 명령어입니다.", inline=False)
        embed.add_field(name="!26/서버등록", value="26이가 당신의 서버를 등록 해드리는 서비스에요!(사실 제작자가 하지만)", inline=False)
        embed.add_field(name="!26/서버리스트", value="26이가 등록한 서버들을 보여줘요(링크 포함)", inline=False)
        embed.add_field(name="!26/로그인", value="자기 프로필로 로그인 합니다!", inline=False)
        embed.set_thumbnail(url="https://cdn.discordapp.com/attachments/689072569076547585/689086114107228186/d201436a554453b7.jpg")
        embed.set_image(url="https://media.giphy.com/media/kDXxjwpazm9b2LzOMc/giphy.gif")
        embed.set_footer(text="봇에 문제가 있으면 !26/고치기 를 해주세요. 그래도 안 되면 𝟚𝟞𝕛𝕡𝟘𝟛#6242 로 친추 거시고 연락 주세요!", icon_url="https://media.giphy.com/media/kDXxjwpazm9b2LzOMc/giphy.gif")
        await message.channel.send("여기 명령어 메뉴판입니다!", embed=embed)

    if message.content.startswith("!26/자기소개"):
        await message.channel.send("```❤️이름: 🍭26이봇🍪``` ```🧡주인: 모두 다(이상한 사람 빼고)``` ```💛취미: 주인이랑 놀기```")

    if message.content.startswith('!영화순위'):
        # http://ticket2.movie.daum.net/movie/movieranklist.aspx
        i1 = 0 # 랭킹 string값
        embed = discord.Embed(
            title = "영화순위",
            description = "영화순위입니다.",
            colour= discord.Color.red()
        )
        hdr = {'User-Agent': 'Mozilla/5.0'}
        url = 'http://ticket2.movie.daum.net/movie/movieranklist.aspx'
        print(url)
        req = Request(url, headers=hdr)
        html = urllib.request.urlopen(req)
        bsObj = bs4.BeautifulSoup(html, "html.parser")
        moviechartBase = bsObj.find('div', {'class': 'main_detail'})
        moviechart1 = moviechartBase.find('ul', {'class': 'list_boxthumb'})
        moviechart2 = moviechart1.find_all('li')

        for i in range(0, 20):
            i1 = i1+1
            stri1 = str(i1) # i1은 영화랭킹을 나타내는데 사용됩니다
            print()
            print(i)
            print()
            moviechartLi1 = moviechart2[i]  # ------------------------- 1등랭킹 영화---------------------------
            moviechartLi1Div = moviechartLi1.find('div', {'class': 'desc_boxthumb'})  # 영화박스 나타내는 Div
            moviechartLi1MovieName1 = moviechartLi1Div.find('strong', {'class': 'tit_join'})
            moviechartLi1MovieName = moviechartLi1MovieName1.text.strip()  # 영화 제목
            print(moviechartLi1MovieName)

            moviechartLi1Ratting1 = moviechartLi1Div.find('div', {'class': 'raking_grade'})
            moviechartLi1Ratting2 = moviechartLi1Ratting1.find('em', {'class': 'emph_grade'})
            moviechartLi1Ratting = moviechartLi1Ratting2.text.strip()  # 영화 평점
            print(moviechartLi1Ratting)

            moviechartLi1openDay1 = moviechartLi1Div.find('dl', {'class': 'list_state'})
            moviechartLi1openDay2 = moviechartLi1openDay1.find_all('dd')  # 개봉날짜, 예매율 두개포함한 dd임
            moviechartLi1openDay3 = moviechartLi1openDay2[0]
            moviechartLi1Yerating1 = moviechartLi1openDay2[1]
            moviechartLi1openDay = moviechartLi1openDay3.text.strip()  # 개봉날짜
            print(moviechartLi1openDay)
            moviechartLi1Yerating = moviechartLi1Yerating1.text.strip()  # 예매율 ,랭킹변동
            print(moviechartLi1Yerating)  # ------------------------- 1등랭킹 영화---------------------------
            print()
            embed.add_field(name='---------------랭킹'+stri1+'위---------------', value='\n영화제목 : '+moviechartLi1MovieName+'\n영화평점 : '+moviechartLi1Ratting+'점'+'\n개봉날짜 : '+moviechartLi1openDay+'\n예매율,랭킹변동 : '+moviechartLi1Yerating, inline=False) # 영화랭킹


        await message.channel.send(embed=embed)


    if message.content.startswith("!급식"):
        embed = discord.Embed(
            title='군포 E 비즈니스 고등학교 급식',
            description='급식입니다.',
            colour=discord.Colour.green()
        )
        embed.add_field(name='오늘', value=급식.lunchtext(), inline=False)
        embed.add_field(name='내일', value=급식.lunchtextD1(), inline=False)
        embed.add_field(name='모래', value=급식.lunchtextD2(), inline=False)
        await message.channel.send(embed=embed)

    
    if message.content.startswith("!복권"):
        Text = ""
        number = [1, 2, 3, 4, 5, 6, 7]
        count = 0
        for i in range(0, 7):
            num = random.randrange(1, 46)
            number[i] = num
            if count >= 1:
                for i2 in range(0, i):
                    if number[i] == number[i2]:  # 만약 현재랜덤값이 이전숫자들과 값이 같다면
                        numberText = number[i]
                        print("작동 이전값 : " + str(numberText))
                        number[i] = random.randrange(1, 46)
                        numberText = number[i]
                        print("작동 현재값 : " + str(numberText))
                        if number[i] == number[i2]:  # 만약 다시 생성한 랜덤값이 이전숫자들과 또 같다면
                            numberText = number[i]
                            print("작동 이전값 : " + str(numberText))
                            number[i] = random.randrange(1, 46)
                            numberText = number[i]
                            print("작동 현재값 : " + str(numberText))
                            if number[i] == number[i2]:  # 만약 다시 생성한 랜덤값이 이전숫자들과 또 같다면
                                numberText = number[i]
                                print("작동 이전값 : " + str(numberText))
                                number[i] = random.randrange(1, 46)
                                numberText = number[i]
                                print("작동 현재값 : " + str(numberText))

            count = count + 1
            Text = Text + "  " + str(number[i])

        print(Text.strip())
        embed = discord.Embed(
            title="복권 숫자!",
            description=Text.strip(),
            colour=discord.Color.red()
        )
        await message.channel.send(embed=embed)

    if message.content.startswith('!검색'):
        Text = ""
        learn = message.content.split(" ")
        vrsize = len(learn)  # 배열크기
        vrsize = int(vrsize)
        for i in range(1, vrsize):  # 띄어쓰기 한 텍스트들 인식함
            Text = Text + " " + learn[i]
        encText = Text

        chromedriver_dir = r'C:\selum\chromedriver_win32\chromedriver.exe' #크롬드라이버 경로
        driver = webdriver.Chrome(chromedriver_dir)
        driver.get('https://www.youtube.com/results?search_query='+encText) #유튜브 검색링크
        source = driver.page_source
        bs = bs4.BeautifulSoup(source, 'lxml')
        entire = bs.find_all('a', {'id': 'video-title'}) # a태그에서 video title 이라는 id를 찾음

        embed = discord.Embed(
            title="영상들!",
            description="검색한 영상 결과",
            colour=discord.Color.blue())

        for i in range(0, 5):
            entireNum = entire[i]
            entireText = entireNum.text.strip()  # 영상제목
            print(entireText)
            test1 = entireNum.get('href')  # 하이퍼링크
            print(test1)
            rink = 'https://www.youtube.com'+test1
           # embed.add_field(name=str(i+1)+'번째 영상',value=entireText + '\n링크 : '+rink)
            embed.add_field(name=str(i + 1) + '번째 영상', value='\n' + '[%s](<%s>)' % (entireText, rink),
                            inline=False)  # [텍스트](<링크>) 형식으로 적으면 텍스트 하이퍼링크 만들어집니다
            searchYoutubeHref[i] = rink
        await message.channel.send(embed=embed)

    if message.content.startswith('1'):

        if not searchYoutubeHref: #저장된 하이퍼링크가 없다면
            print('searchYoutubeHref 안에 값이 존재하지 않습니다.')
            await client.send_message(message.channel, embed=discord.Embed(description="검색한 영상이 없습니다."))
        else:
            print(searchYoutubeHref[0])
            server = message.server
            voice_client = client.voice_client_in(server)
            url = searchYoutubeHref[0]
            player = await voice_client.create_ytdl_player(url, after=lambda: check_queue(server.id))
            print(player.is_playing())
            players[server.id] = player
            await message.channel.send(embed=discord.Embed(description="재생한다!!!!"))
            print(player.is_playing())
            player.start()

            for i in range(0,5):
                del searchYoutubeHref[i]

    if message.content.startswith('2'):

        if not searchYoutubeHref:
            print('searchYoutubeHref 안에 값이 존재하지 않습니다.')
            await message.channel.send(embed=discord.Embed(description="검색한 영상이 없습니다."))
        else:
            print(searchYoutubeHref[1])
            server = message.server
            voice_client = client.voice_client_in(server)
            url = searchYoutubeHref[1]
            player = await voice_client.create_ytdl_player(url, after=lambda: check_queue(server.id))
            print(player.is_playing())
            players[server.id] = player
            await message.channel.send(embed=discord.Embed(description="재생한다!!!!"))
            print(player.is_playing())
            player.start()

            for i in range(0,5):
                del searchYoutubeHref[i]

    if message.content.startswith('3'):

        if not searchYoutubeHref:
            print('searchYoutubeHref 안에 값이 존재하지 않습니다.')
            await client.send_message(message.channel, embed=discord.Embed(description="검색한 영상이 없습니다."))
        else:
            print(searchYoutubeHref[2])
            server = message.server
            voice_client = client.voice_client_in(server)
            url = searchYoutubeHref[2]
            player = await voice_client.create_ytdl_player(url, after=lambda: check_queue(server.id))
            print(player.is_playing())
            players[server.id] = player
            await message.channel.send(embed=discord.Embed(description="재생한다!!!!"))
            print(player.is_playing())
            player.start()

            for i in range(0,5):
                del searchYoutubeHref[i]

    if message.content.startswith('4'):

        if not searchYoutubeHref:
            print('searchYoutubeHref 안에 값이 존재하지 않습니다.')
            await message.channel.send(embed=discord.Embed(description="검색한 영상이 없습니다."))
        else:
            print(searchYoutubeHref[3])
            server = message.server
            voice_client = client.voice_client_in(server)
            url = searchYoutubeHref[3]
            player = await voice_client.create_ytdl_player(url, after=lambda: check_queue(server.id))
            print(player.is_playing())
            players[server.id] = player
            await message.channel.send(embed=discord.Embed(description="재생한다!!!!"))
            print(player.is_playing())
            player.start()

            for i in range(0,5):
                del searchYoutubeHref[i]

    if message.content.startswith('5'):

        if not searchYoutubeHref:
            print('searchYoutubeHref 안에 값이 존재하지 않습니다.')
            await client.send_message(message.channel, embed=discord.Embed(description="검색한 영상이 없습니다."))
        else:
            print(searchYoutubeHref[4])
            server = message.server
            voice_client = client.voice_client_in(server)
            url = searchYoutubeHref[4]
            player = await voice_client.create_ytdl_player(url, after=lambda: check_queue(server.id))
            print(player.is_playing())
            players[server.id] = player
            await message.channel.send(embed=discord.Embed(description="재생한다!!!!"))
            print(player.is_playing())
            player.start()

            for i in range(0,5):
                del searchYoutubeHref[i]


    if message.content.startswith('!이모티콘'):

        emoji = [" ꒰⑅ᵕ༚ᵕ꒱ ", " ꒰◍ˊ◡ˋ꒱ ", " ⁽⁽◝꒰ ˙ ꒳ ˙ ꒱◜⁾⁾ ", " ༼ つ ◕_◕ ༽つ ", " ⋌༼ •̀ ⌂ •́ ༽⋋ ",
                 " ( ･ิᴥ･ิ) ", " •ө• ", " ค^•ﻌ•^ค ", " つ╹㉦╹)つ ", " ◕ܫ◕ ", " ᶘ ͡°ᴥ͡°ᶅ ", " ( ؕؔʘ̥̥̥̥ ه ؔؕʘ̥̥̥̥ ) ",
                 " ( •́ ̯•̀ ) ",
                 " •̀.̫•́✧ ", " '͡•_'͡• ", " (΄◞ิ౪◟ิ‵) ", " ˵¯͒ བ¯͒˵ ", " ͡° ͜ʖ ͡° ", " ͡~ ͜ʖ ͡° ", " (づ｡◕‿‿◕｡)づ ",
                 " ´_ゝ` ", " ٩(͡◕_͡◕ ", " ⁄(⁄ ⁄•⁄ω⁄•⁄ ⁄)⁄ ", " ٩(͡ï_͡ï☂ ", " ௐ ", " (´･ʖ̫･`) ", " ε⌯(ง ˙ω˙)ว ",
                 " (っ˘ڡ˘ς) ", "●▅▇█▇▆▅▄▇", "╋╋◀", "︻╦̵̵̿╤──", "ー═┻┳︻▄", "︻╦̵̵͇̿̿̿̿══╤─",
                 " ጿ ኈ ቼ ዽ ጿ ኈ ቼ ዽ ጿ ", "∑◙█▇▆▅▄▃▂", " ♋♉♋ ", " (๑╹ω╹๑) ", " (╯°□°）╯︵ ┻━┻ ",
                 " (///▽///) ", " σ(oдolll) ", " 【o´ﾟ□ﾟ`o】 ", " ＼(^o^)／ ", " (◕‿‿◕｡) ", " ･ᴥ･ ", " ꈍ﹃ꈍ "
                                                                                                 " ˃̣̣̣̣̣̣︿˂̣̣̣̣̣̣ ",
                 " ( ◍•㉦•◍ ) ", " (｡ì_í｡) ", " (╭•̀ﮧ •́╮) ", " ଘ(੭*ˊᵕˋ)੭ ", " ´_ゝ` ", " (~˘▾˘)~ "] # 이모티콘 배열입니다.

        randomNum = random.randrange(0, len(emoji)) # 0 ~ 이모티콘 배열 크기 중 랜덤숫자를 지정합니다.
        print("랜덤수 값 :" + str(randomNum))
        print(emoji[randomNum])
        await message.channel.send(embed=discord.Embed(description=emoji[randomNum])) # 랜덤 이모티콘을 메시지로 출력합니다.

    if message.content.startswith('!주사위'):

        randomNum = random.randrange(1, 7) # 1~6까지 랜덤수
        print(randomNum)
        if randomNum == 1:
            await message.channel.send(embed=discord.Embed(description=':game_die: '+ ':one:'))
        if randomNum == 2:
            await message.channel.send(embed=discord.Embed(description=':game_die: ' + ':two:'))
        if randomNum ==3:
            await message.channel.send(embed=discord.Embed(description=':game_die: ' + ':three:'))
        if randomNum ==4:
            await message.channel.send(embed=discord.Embed(description=':game_die: ' + ':four:'))
        if randomNum ==5:
            await message.channel.send(embed=discord.Embed(description=':game_die: ' + ':five:'))
        if randomNum ==6:
            await message.channel.send(embed=discord.Embed(description=':game_die: ' + ':six: '))

    if message.content.startswith('!타이머'):

        Text = ""
        learn = message.content.split(" ")
        vrsize = len(learn)  # 배열크기
        vrsize = int(vrsize)
        for i in range(1, vrsize):  # 띄어쓰기 한 텍스트들 인식함
            Text = Text + " " + learn[i]

        secint = int(Text)
        sec = secint

        for i in range(sec, 0, -1):
            print(i)
            await message.channel.send(embed=discord.Embed(description='타이머 작동중 : '+str(i)+'초'))
            time.sleep(1)

        else:
            print("땡")
            await message.channel.send(embed=discord.Embed(description='타이머 종료'))

    if message.content.startswith('!제비뽑기'):
        channel = message.channel
        embed = discord.Embed(
            title='제비뽑기',
            description='각 번호별로 번호를 지정합니다.',
            colour=discord.Colour.blue()
        )

        embed.set_footer(text='끗')


        Text = ""
        learn = message.content.split(" ")
        vrsize = len(learn)  # 배열크기
        vrsize = int(vrsize)
        for i in range(1, vrsize):  # 띄어쓰기 한 텍스트들 인식함
            Text = Text + " " + learn[i]
        print(Text.strip()) #입력한 명령어

        number = int(Text)

        List = []
        num = random.randrange(0, number)
        for i in range(number):
            while num in List:  # 중복일때만
                num = random.randrange(0, number)  # 다시 랜덤수 생성

            List.append(num)  # 중복 아닐때만 리스트에 추가
            embed.add_field(name=str(i+1) + '번째', value=str(num+1), inline=True)

        print(List)
        await message.channel.send(embed=embed)

    if message.content.startswith('!이미지'):

        Text = ""
        learn = message.content.split(" ")
        vrsize = len(learn)  # 배열크기
        vrsize = int(vrsize)
        for i in range(1, vrsize):  # 띄어쓰기 한 텍스트들 인식함
            Text = Text + " " + learn[i]
        print(Text.strip())  # 입력한 명령어

        randomNum = random.randrange(0, 40) # 랜덤 이미지 숫자

        location = Text
        enc_location = urllib.parse.quote(location) # 한글을 url에 사용하게끔 형식을 바꿔줍니다. 그냥 한글로 쓰면 실행이 안됩니다.
        hdr = {'User-Agent': 'Mozilla/5.0'}
        # 크롤링 하는데 있어서 가끔씩 안되는 사이트가 있습니다.
        # 그 이유는 사이트가 접속하는 상대를 봇으로 인식하였기 때문인데
        # 이 코드는 자신이 봇이 아닌것을 증명하여 사이트에 접속이 가능해집니다!
        url = 'https://search.naver.com/search.naver?where=image&sm=tab_jum&query=' + enc_location # 이미지 검색링크+검색할 키워드
        print(url)
        req = Request(url, headers=hdr)
        html = urllib.request.urlopen(req)
        bsObj = bs4.BeautifulSoup(html, "html.parser") # 전체 html 코드를 가져옵니다.
        # print(bsObj)
        imgfind1 = bsObj.find('div', {'class': 'photo_grid _box'}) # bsjObj에서 div class : photo_grid_box 의 코드를 가져옵니다.
        # print(imgfind1)
        imgfind2 = imgfind1.findAll('a', {'class': 'thumb _thumb'}) # imgfind1 에서 모든 a태그 코드를 가져옵니다.
        imgfind3 = imgfind2[randomNum]  # 0이면 1번째사진 1이면 2번째사진 형식으로 하나의 사진 코드만 가져옵니다.
        imgfind4 = imgfind3.find('img') # imgfind3 에서 img코드만 가져옵니다.
        imgsrc = imgfind4.get('data-source') # imgfind4 에서 data-source(사진링크) 의 값만 가져옵니다.
        print(imgsrc)
        embed = discord.Embed(
            colour=discord.Colour.green()
        )
        embed.add_field(name='검색 : '+Text, value='링크 : '+imgsrc, inline=False)
        embed.set_image(url=imgsrc) # 이미지의 링크를 지정해 이미지를 설정합니다.
        await message.channel.send(embed=embed) # 메시지를 보냅니다.


    if message.content.startswith('!members'):
        x = message.server.members
        for member in x:
            print(member.name)  # you'll just print out Member objects your way. 

acess_token = os.environ["BOT_TOKEN"]
client.run(access_token)
